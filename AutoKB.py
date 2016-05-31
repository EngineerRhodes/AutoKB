import json
import requests
import sys
import datetime
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font

#Create work book file
wb = Workbook()
wb.save('HelpDesk Article Review Sheet.xlsx')

#Make first worksheet active
ws = wb.active
  
#Create a bold font 
bold_it = Font(name='Calibri', size=11, bold=True)

#Bold things
a1 = ws['A1']
b1 = ws['B1']
c1 = ws['C1']
d1 = ws['C1']

#Headers
ws['A1'] = 'Article Title - click title to access article'
ws['B1'] = 'Last Updated Date'
ws['C1'] = 'Created Date'
ws['D1'] = 'Notes'

a1.font = bold_it
b1.font = bold_it
c1.font = bold_it
d1.font = bold_it
# Just a temporary keeper
results_pages = []

# The array that we actually want to populate
articles = []

#Zendesk help site
url ='https://huddle.zendesk.com/api/v2/help_center/articles.json?page='

page_number = 1


#Search through paginated Json data to gether all help site article data; generates as a list of 
while True:
  response1 = requests.get(url + str(page_number))
  json_data = json.loads(response1.text)
  results_pages.append(json_data)
  if json_data.get('next_page') == None:
    break
  page_number = page_number + 1
  
#Grab needed keys from compiled JSON List of Dictionaries
for page in results_pages:  
  for article in page['articles']:
    
    
    article_to_append = {
      'title': article['title'],
      'html_url': article['html_url'],
      'updated_at': article['updated_at'],
      'created_at': article['created_at']}
    
    articles.append(article_to_append)


articles.sort(key=lambda x: datetime.datetime.strptime(x['updated_at'], "%Y-%m-%dT%H:%M:%SZ"))


#Counter starts at 2 to account for Titles in first row
counter = 2

#Go through articles, grab there data, output into excel spreadsheet
for entry in articles:
 
  title = entry['title']  
  url = entry['html_url']  
  updated_at = entry['updated_at'][0:10]
  created_at = entry['created_at'][0:10]

  current_cell_A = 'A' + str(counter)
  current_cell_B = 'B' + str(counter)
  current_cell_C = 'C' + str(counter)
  
#These are likely superfluos, but add clarity to what is happening  
  column_A = title
  column_B = updated_at
  column_hA = url
  column_C = created_at

  output_list = load_workbook('HelpDesk Article Review Sheet.xlsx', True)
  current_worksheet = output_list.active
  
  

#Write values to cells
  ws.cell(current_cell_A).value = column_A
  ws.cell(current_cell_A).hyperlink = column_hA
  ws.cell(current_cell_B).value = column_B
  ws.cell(current_cell_C).value = column_C
  
  counter = counter + 1
      
wb.save('HelpDesk Article Review Sheet.xlsx')
 
